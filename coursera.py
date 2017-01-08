import requests
import json
import random
import argparse
from bs4 import BeautifulSoup
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


XLS_HEADERFILL = PatternFill(start_color='E1E4D2', end_color='E1E4D2', fill_type='solid')
XLS_TITLEFONT = Font(size=11, bold=True, color='000000')


def get_console_arguments():
    parser = argparse.ArgumentParser(description='Gets random courses from coursera and outputs them to xlsx')
    parser.add_argument('--s', '--save_as', dest='save_as', default='result.xlsx',
                        help='save the output file as... (default: result.xlsx)')

    args = parser.parse_args()
    return args


def get_random_courses(amount=20):
    base_url = 'https://www.coursera.org/sitemap~www~courses.xml'
    html_data = requests.get(base_url)
    root = etree.fromstring(html_data.content)
    name_spaces = {'urls': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
    urls_obj_list = root.xpath("//urls:loc", namespaces=name_spaces)
    all_urls_list = [url.text for url in urls_obj_list]
    random_urls = random.sample(all_urls_list, amount)
    return random_urls


def get_page(url):
    response = requests.get(url)
    page = response.text
    soup = BeautifulSoup(page, 'lxml')
    return soup


def get_course_language(soup):
    class_table = "basic-info-table bt3-table bt3-table-striped bt3-table-bordered bt3-table-responsive"
    table = soup.find('table', class_=class_table)
    if table:
        row_info = []
        for row in table.find_all("tr")[:3]:
            row_info.extend([cell.get_text(strip=True)
                             for cell in row.find_all("td")])
        try:
            language_index = row_info.index('Language') + 1
            language = row_info[language_index]
        except ValueError:
            language = None
        return language


def get_start_date(soup):
    try:
        data_from_script = soup.select(
            'script[type="application/ld+json"]')[0].text
        data_json = json.loads(data_from_script)
        start_date = data_json['hasCourseInstance'][0]['startDate']
        return start_date
    except IndexError:
        return None


def get_course_rating(soup):
    rating = soup.find('div', class_='ratings-text bt3-visible-xs')
    return None if rating is None else rating.text


def get_course_title(soup):
    title = soup.find('div', class_='title display-3-text')
    return None if title is None else title.text


def get_course_duration(soup):
    duration = soup.findAll('div', class_='week-heading body-2-text')
    return len(duration)


def get_course_data(course_slug):
    soup = get_page(course_slug)
    return {'title': get_course_title(soup),
            'startdate': get_start_date(soup),
            'rating': get_course_rating(soup),
            'duration': get_course_duration(soup),
            'language': get_course_language(soup),
            'url': course_slug}


def compile_workbook(all_courses_info):
    wb = Workbook()
    sheet = wb.active
    sheet.append(['Title', 'Start date', 'Rating',
                  'Duration in weeks', 'Language', 'Url'])

    for row in sheet["A1:F1"]:
        for cell in row:
            cell.fill = XLS_HEADERFILL
            cell.font = XLS_TITLEFONT

    for course in all_courses_info:
        if course['language'] is None:
            course['language'] = 'Unknown'
        elif course['rating'] is None:
            course['rating'] = 'Unknown'

        course_info = [course['title'], course['startdate'], course['rating'],
                       course['duration'], course['language'], course['url']]
        sheet.append(course_info)

    return wb


def write_excel(courses_workbook, output_file):
    courses_workbook.save(output_file)


def main():
    args = get_console_arguments()
    output_file = args.save_as
    all_courses_info = [get_course_data(url) for url in get_random_courses()]
    courses_workbook = compile_workbook(all_courses_info)
    write_excel(courses_workbook, output_file)

    print('the file has been successfully saved')


if __name__ == '__main__':
    main()
